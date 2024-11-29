"""
Generates a 1-D nerve path mesh.
"""

from __future__ import division

import json
import openpyxl

from cmlibs.utils.zinc.field import findOrCreateFieldStoredMeshLocation
from cmlibs.zinc.element import Element, Elementbasis
from cmlibs.zinc.field import Field
from cmlibs.zinc.node import Node
from scaffoldmaker.annotation.annotationgroup import AnnotationGroup
from scaffoldmaker.meshtypes.scaffold_base import Scaffold_base
from cmlibs.utils.zinc.field import findOrCreateFieldCoordinates, findOrCreateFieldStoredString, findOrCreateFieldGroup

class MeshType_1d_nervepath1(Scaffold_base):
    '''
    classdocs
    '''
    @staticmethod
    def getName():
        return '1D Nerve Path 1'

    @classmethod
    def generateBaseMesh(cls, region, options):
        """
        :param region: Zinc region to define model in. Must be empty.
        :param options: Dict containing options. See getDefaultOptions().
        :return: [] empty list of AnnotationGroup, None
        """
        coordinateDimensions = 3

        fieldmodule = region.getFieldmodule()
        coordinates = findOrCreateFieldCoordinates(fieldmodule, components_count=coordinateDimensions)
        cache = fieldmodule.createFieldcache()

        nodes = fieldmodule.findNodesetByFieldDomainType(Field.DOMAIN_TYPE_NODES)
        nodetemplate = nodes.createNodetemplate()
        nodetemplate.defineField(coordinates)
        nodetemplate.setValueNumberOfVersions(coordinates, -1, Node.VALUE_LABEL_VALUE, 1)

        nodeIdentifier = 1

        mesh = fieldmodule.findMeshByDimension(1)
        linearBasis = fieldmodule.createElementbasis(1, Elementbasis.FUNCTION_TYPE_LINEAR_LAGRANGE)
        eft = mesh.createElementfieldtemplate(linearBasis)
        elementtemplate = mesh.createElementtemplate()
        elementtemplate.setElementShapeType(Element.SHAPE_TYPE_LINE)
        result = elementtemplate.defineField(coordinates, -1, eft)
        elementIdentifier = 1

        filename1 = "C:\\Users\\mlin865\\map\\packages\\sparc\\scaffoldmaker\\inputFile\\nerveAnnotations_manInBox.json"
        excelFile = "C:\\Users\\mlin865\\map\\packages\\sparc\\scaffoldmaker\\inputFile\\M2.6 Origin, destination and components for major nerve pathways for 3D whole-body.xlsx"

        # Extract identifiers
        wb = openpyxl.load_workbook(excelFile)
        ws = wb['Sheet1']
        namesInMIB = []
        termIDs = []
        for i in range(1, ws.max_row):
            if ws.cell(row=i, column=1).value == 'Yes':
                branchName = ws.cell(row=i, column=2).value
                # remove second name
                namesInMIB.append(branchName.split(" (")[0])  # parse to get term used in man-in-box
                contentFromIDColumn = ws.cell(row=i, column=3).value
                if contentFromIDColumn == "REQUESTED":
                    termIDs.append('None')
                else:
                    termIDs.append(ws.cell(row=i, column=3).value)


        marker_data = {}
        marker_data_group = {}
        with open(filename1) as json_data:
            d1 = json.load(json_data)
            json_data.close()

        markerCount = 0
        for feature in d1:
            marker_name = feature['group']
            marker_point = feature['feature']['geometry']['coordinates'][0]
            marker_group = feature['region'].replace('__annotation/', '')

            if marker_group in marker_data_group.keys():
                marker_data_group[marker_group].append(marker_name)
            else:
                marker_data_group[marker_group] = [marker_name]

            if marker_name in marker_data.keys():
                pass
                # print(marker_name, marker_group)
            else:
                marker_data[marker_name] = marker_point
                markerCount += 1
                # print('New marker in', marker_group, ': ', marker_name)

        # marker coordinates
        cleanGroupList = []
        cleanMarkerList = []
        cleanMarkerNodeList = []
        cleanMarkerCoordList = []
        allAnnotationGroups = []
        elementIdsList = [[] for n in range(markerCount)]
        xiList = [[] for n in range(markerCount)]

        print('Number of branches = ', len(marker_data_group.keys()))
        # print('Number of markers = ', markerCount)

        for marker_group in marker_data_group.keys():
            origins = []
            waypoints = []
            waypointsOrder = []
            destinations = []

            # Check for duplicate groups
            groupNotInList = 1
            for group in cleanGroupList:
                if marker_group == group:
                    print('Duplicate groups', marker_group)
                    groupNotInList = 0
                    break
            if groupNotInList:
                cleanGroupList.append(marker_group)

                # Create annotation group
                # Find branch name identifiers
                branchID = []
                for i in range(len(namesInMIB)):
                    if marker_group == namesInMIB[i]:
                        branchID = termIDs[i]
                if branchID:
                    pass
                else:
                    print('Branch has no ID -', marker_group)
                branchGroup = AnnotationGroup(region, (marker_group, branchID if branchID else 'None'))

            # if marker_group in ["Left obturator nerve"]:
            # Discover markers in each group (duplicate included)
            for name in marker_data_group[marker_group]:
                # Separate names and tags for origin/destination/waypoint
                splitResult = name.split("(")
                currentMarkerName = splitResult[0]
                if len(splitResult) > 1:
                    markerTag = splitResult[1].split(")")[0]
                else:
                    print('Markers with no tags -', currentMarkerName)

                xyz = marker_data[name] # Get coordinates

                # check if point exists before
                markerNotInList = 1
                for i in range(len(cleanMarkerList)):
                    marker = cleanMarkerList[i]
                    if currentMarkerName == marker:
                        markerNotInList = 0
                        currentNodeId = cleanMarkerNodeList[i]
                        break
                if markerNotInList: # if new point, create a node
                    cleanMarkerList.append(currentMarkerName)
                    cleanMarkerNodeList.append(nodeIdentifier)
                    cleanMarkerCoordList.append(xyz)
                    node = nodes.createNode(nodeIdentifier, nodetemplate)
                    cache.setNode(node)
                    coordinates.setNodeParameters(cache, -1, Node.VALUE_LABEL_VALUE, 1, xyz)
                    # print('Node', nodeIdentifier, name)
                    # if marker_group == 'Left radial nerve':
                    #     print('Node', nodeIdentifier, name)
                    currentNodeId = nodeIdentifier
                    nodeIdentifier += 1

                # store tag identity
                if 'origin' in markerTag:
                    origins.append(currentNodeId)
                elif 'destination' in markerTag:
                    destinations.append(currentNodeId)
                elif 'waypoint' in markerTag:
                    waypoints.append(currentNodeId)
                    waypointsOrder.append(markerTag)
                else:
                    print('Unidentified marker tag - ', markerTag)

            # Remove duplicate tags
            origins = list(set(origins))

            waypointsClean = []
            waypointsOrderClean = []
            for i in range(len(waypoints)):
                if waypoints[i] not in waypointsClean:
                    waypointsClean.append(waypoints[i])
                    waypointsOrderClean.append(waypointsOrder[i])

            destinations = list(set(destinations))

            # print('Branch', marker_group)
            # print('Origins', origins)
            # print('Waypoints', waypoints)
            # print('Destinations', destinations)

            # Re-arrange points in each tag from cranial to caudal
            if len(waypoints) > 1:
            #     if marker_group == 'Left radial nerve':
            #         debug = 1
                waypoints = sortPointsUsingTag(waypointsClean, waypointsOrderClean)

            # create elements to connect points up when all points are processed in the group
            if len(waypoints) == 0 and len(origins) == 1 and len(destinations) > 0:
                allAnnotationGroups.append(branchGroup)
                for d in range(len(destinations)):
                    element = mesh.createElement(elementIdentifier, elementtemplate)
                    nIds = [origins[0], destinations[d]]
                    element.setNodesByIdentifier(eft, nIds)
                    elementIdsList, xiList = findElementAndXi(nIds, elementIdentifier, elementIdsList, xiList)
                    elementIdentifier = elementIdentifier + 1
                    meshGroup = branchGroup.getMeshGroup(mesh)
                    meshGroup.addElement(element)
            elif len(waypoints) == 0 and len(origins) > 1 and len(destinations) == 1:
                allAnnotationGroups.append(branchGroup)
                for o in range(len(origins)):
                    element = mesh.createElement(elementIdentifier, elementtemplate)
                    nIds = [origins[o], destinations[0]]
                    element.setNodesByIdentifier(eft, nIds)
                    elementIdsList, xiList = findElementAndXi(nIds, elementIdentifier, elementIdsList, xiList)
                    elementIdentifier = elementIdentifier + 1
                    meshGroup = branchGroup.getMeshGroup(mesh)
                    meshGroup.addElement(element)

            elif len(waypoints) > 0:
                allAnnotationGroups.append(branchGroup)
                if len(origins) > 0:
                    for o in range(len(origins)):
                        element = mesh.createElement(elementIdentifier, elementtemplate)
                        nIds = [origins[o], waypoints[0]]
                        element.setNodesByIdentifier(eft, nIds)
                        elementIdsList, xiList = findElementAndXi(nIds, elementIdentifier, elementIdsList, xiList)
                        # if marker_group == 'Left radial nerve':
                        #     print("A", elementIdentifier, nIds)
                        elementIdentifier = elementIdentifier + 1
                        meshGroup = branchGroup.getMeshGroup(mesh)
                        meshGroup.addElement(element)
                for w in range(len(waypoints) - 1):
                    element = mesh.createElement(elementIdentifier, elementtemplate)
                    nIds = [waypoints[w], waypoints[w + 1]]
                    element.setNodesByIdentifier(eft, nIds)
                    elementIdsList, xiList = findElementAndXi(nIds, elementIdentifier, elementIdsList, xiList)
                    # if marker_group == 'Left radial nerve':
                    #     print("B", elementIdentifier, nIds)
                    elementIdentifier = elementIdentifier + 1
                    meshGroup = branchGroup.getMeshGroup(mesh)
                    meshGroup.addElement(element)
                if len(destinations) > 0:
                    for d in range(len(destinations)):
                        element = mesh.createElement(elementIdentifier, elementtemplate)
                        nIds = [waypoints[-1], destinations[d]]
                        element.setNodesByIdentifier(eft, nIds)
                        elementIdsList, xiList = findElementAndXi(nIds, elementIdentifier, elementIdsList, xiList)
                        # print("C", elementIdentifier, nIds)
                        elementIdentifier = elementIdentifier + 1
                        meshGroup = branchGroup.getMeshGroup(mesh)
                        meshGroup.addElement(element)

            elif len(origins) > 0 and len(waypoints) < 1 and len(destinations) < 1:
                print('Invalid nerve with only origin -', marker_group, len(origins), len(waypoints), len(destinations))
            elif len(origins) < 1 and len(waypoints) < 1 and len(destinations) > 0:
                print('Invalid nerve with only destination -', marker_group, len(origins), len(waypoints), len(destinations))
            elif len(origins) < 1 and len(waypoints) > 0 and len(destinations) < 1:
                print('Invalid nerve with only waypoint -', marker_group, len(origins), len(waypoints), len(destinations))
            else:
                print('Code could not deal with this -', marker_group, len(origins), len(waypoints), len(destinations))


        # Make markers
        markerGroup = findOrCreateFieldGroup(fieldmodule, "marker")
        markerName = findOrCreateFieldStoredString(fieldmodule, name="marker_name")
        markerLocation = findOrCreateFieldStoredMeshLocation(fieldmodule, mesh, name="marker_location")
        markerCoordinates = findOrCreateFieldCoordinates(fieldmodule, name="marker coordinates").castFiniteElement()

        markerPoints = markerGroup.getOrCreateNodesetGroup(nodes)
        markerTemplateInternal = nodes.createNodetemplate()
        markerTemplateInternal.defineField(markerCoordinates)
        markerTemplateInternal.defineField(markerName)
        markerTemplateInternal.defineField(markerLocation)

        for i in range(len(cleanMarkerList)):
            markerPoint = markerPoints.createNode(nodeIdentifier, markerTemplateInternal)
            nodeIdentifier += 1
            cache.setNode(markerPoint)
            markerCoordinates.setNodeParameters(cache, -1, Node.VALUE_LABEL_VALUE, 1, cleanMarkerCoordList[i])
            markerName.assignString(cache, cleanMarkerList[i])
            elementID = elementIdsList[i + 1]
            xi = xiList[i + 1]
            element = mesh.findElementByIdentifier(elementID)
            markerLocation.assignMeshLocation(cache, element, xi)

        # # add groups in argon viewer graphics settings
        # argonSettingsFile = "C:\\Users\\mlin865\\map\\workflows\\datasets\\manInBox-Nerves\\Argon_Viewer-previous-docs\\document-dc3a2848.json"
        # modFile = "C:\\Users\\mlin865\\map\\workflows\\datasets\\manInBox-Nerves\\Argon_Viewer-previous-docs\\document-dc3a2848_modified.json"
        #
        # with open(argonSettingsFile) as json_data:
        #     d = json.load(json_data)
        #     json_data.close()
        #
        # graphics = d['RootRegion']['ChildRegions'][1]['Scene']['Graphics']
        # template = copy.deepcopy(d['RootRegion']['ChildRegions'][1]['Scene']['Graphics'][0])
        #
        # for group in allAnnotationGroups:
        #     template = copy.deepcopy(d['RootRegion']['ChildRegions'][1]['Scene']['Graphics'][0])
        #     template['SubgroupField'] = group.getName()
        #     graphics.append(template)
        #
        # with open(modFile, 'w') as f:
        #     json.dump(d, f)

        # # # # Extract groups of nerves from annotation file for Jen
        # reducedFile = "C:\\Users\\mlin865\\manInBox\\files from Jen\\leftVagusNerves.json"
        # extractNerveList = ['Left meningeal branch of left vagus nerve',
        #                     'Left branch between vagus nerve and glossopharyngeal nerve',
        #                     'Left auricular branch of left vagus nerve',
        #                     'Communicating branch of auricular branch of left vagus nerve with left facial nerve',
        #                     'Communicating branch of auricular branch of left vagus nerve with left posterior auricular nerve',
        #                     'Left pharyngeal branch of left vagus nerve to pharyngeal nerve plexus',
        #                     'Lingual branch of left vagus nerve',
        #                     'Left pharyngeal branch of left vagus nerve to superior cervical ganglion',
        #                     'Branch of left vagus nerve to carotid body',
        #                     'Left branch between vagus nerve and superior cervical ganglion',
        #                     'Left superior laryngeal nerve',
        #                     'Left internal laryngeal nerve',
        #                     'Left external laryngeal nerve',
        #                     'Superior branch of left internal laryngeal nerve',
        #                     'Middle branch of left internal laryngeal nerve',
        #                     'Inferior branch of left internal laryngeal nerve',
        #                     'Communicating branch of left internal laryngeal nerve with left recurrent laryngeal nerve',
        #                     'Communicating branch of left external laryngeal nerve with left superior cardiac nerve',
        #                     'Left A cervical cardiopulmonary branch of vagus nerve',
        #                     'Left B cervical cardiopulmonary branch of vagus nerve',
        #                     'Left pulmonary branch A of the vagus nerve',
        #                     'Left pulmonary branch B of the vagus nerve',
        #                     'Left pulmonary branch C of the vagus nerve',
        #                     'Left pulmonary branch D of the vagus nerve',
        #                     'Superior cervical cardiac branch of left vagus nerve',
        #                     'Left recurrent laryngeal nerve',
        #                     'Extra laryngeal branch of left recurrent laryngeal nerve to larynx',
        #                     'Branch of left recurrent laryngeal nerve to muscle of larynx',
        #                     'Esophageal branch of left recurrent laryngeal nerve',
        #                     'Tracheal branch of left recurrent laryngeal nerve',
        #                     'Anterior branch of left recurrent laryngeal nerve',
        #                     'Posterior branch of left recurrent laryngeal nerve',
        #                     'Inferior cervical cardiac branch of left recurrent laryngeal nerve',
        #                     'Inferior cervical cardiac branch of left vagus nerve',
        #                     'Thoracic cardiac branch of left vagus nerve',
        #                     'Cardiac branch of left vagus to deep cardiac plexus',
        #                     'Left branch of left vagus nerve to esophageal nerve plexus',
        #                     'Celiac branch of posterior vagal trunk',
        #                     'Greater posterior gastric nerve',
        #                     'Pyloric branch of greater posterior gastric nerve']
        #
        # # extractNerveList = ['Right vagus X nerve trunk',
        #                        # 'Right meningeal branch of right vagus nerve',
        # #                     'Right branch between vagus nerve and glossopharyngeal nerve',
        # #                     'Right auricular branch of right vagus nerve',
        # #                     'Communicating branch of auricular branch of right vagus nerve with right facial nerve',
        # #                     'Communicating branch of auricular branch of right vagus nerve with right posterior auricular nerve',
        # #                     'Right pharyngeal branch of right vagus nerve to pharyngeal nerve plexus',
        # #                     'Lingual branch of right vagus nerve',
        # #                     'Right pharyngeal branch of right vagus nerve to superior cervical ganglion',
        # #                     'Branch of right vagus nerve to carotid body',
        # #                     'Right branch between vagus nerve and superior cervical ganglion',
        # #                     'Right superior laryngeal nerve',
        # #                     'Right internal laryngeal nerve',
        # #                     'Right external laryngeal nerve',
        # #                     'Superior branch of right internal laryngeal nerve',
        # #                     'Middle branch of right internal laryngeal nerve',
        # #                     'Inferior branch of right internal laryngeal nerve',
        # #                     'Communicating branch of right internal laryngeal nerve with right recurrent laryngeal nerve',
        # #                     'Communicating branch of right external laryngeal nerve with right superior cardiac nerve',
        # #                     'Right A cervical cardiopulmonary branch of vagus nerve',
        # #                     'Right B cervical cardiopulmonary branch of vagus nerve',
        # #                     'Right pulmonary branch A of the vagus nerve',
        # #                     'Right pulmonary branch B of the vagus nerve',
        # #                     'Right pulmonary branch C of the vagus nerve',
        # #                     'Right pulmonary branch D of the vagus nerve',
        # #                     'Right pulmonary branch E of the vagus nerve',
        # #                     'Superior cervical cardiac branch of right vagus nerve',
        # #                     'Right recurrent laryngeal nerve',
        # #                     'Extra laryngeal branch of right recurrent laryngeal nerve to larynx',
        # #                     'Branch of right recurrent laryngeal nerve to muscle of larynx',
        # #                     'Esophageal branch of right recurrent laryngeal nerve',
        # #                     'Tracheal branch of right recurrent laryngeal nerve',
        # #                     'Anterior branch of right recurrent laryngeal nerve',
        # #                     'Posterior branch of right recurrent laryngeal nerve',
        # #                     'Inferior cervical cardiac branch of right recurrent laryngeal nerve',
        # #                     'Inferior cervical cardiac branch of right vagus nerve',
        # #                     'Thoracic cardiac branch of right vagus nerve',
        # #                     'Cardiac branch of right vagus to deep cardiac plexus',
        # #                     'Branch of greater anterior gastric nerve to coeliac nerve plexus',
        # #                     'Greater anterior gastric nerve',
        # #                     'Hepatic branch of anterior vagal trunk',
        # #                     'Right branch of right vagus nerve to esophageal nerve plexus'
        # #                     ]
        # reducedList = []
        # for feature in d1:
        #     marker_group = feature['region'].replace('__annotation/', '')
        #     # if 'intercostal' in marker_group:
        #     if marker_group in extractNerveList:
        #         reducedList.append(feature)
        #
        # with open(reducedFile, 'w') as f:
        #     json.dump(reducedList, f)


        # print('Number of annotated branches', len(allAnnotationGroups))

        return allAnnotationGroups, None

def sortPointsUsingTag(inputTag, tagOrder, debug=0):
    """
    Sorts origins, waypoints and destinations in ascending order based on the number appended to the tag.
    :param inputTag: List of node identifiers for points to be sorted.
    :param tagOrder: List of order corresponding to the nodes in inputTag.
    :param debug: True to enter debug mode.
    :return list of node identifiers sorted in ascending order.
    """

    if debug:
        print('inputTag', inputTag)
    sortedMarkers = []
    li = []
    for i in range(len(tagOrder)):
        order = int(tagOrder[i].split(" ")[1])
        li.append([order, i])
    if debug:
        print('li =', li)
    li.sort()
    if debug:
        print('newOrder', li)
    sort_index = []
    for x in li:
        sort_index.append(x[1])
    if debug:
        print('sort_index', sort_index)
    for index in sort_index:
        sortedMarkers.append(inputTag[index])
    if debug:
        print('sortedMarkers', sortedMarkers)

    return sortedMarkers

def findElementAndXi(nIds, elementIdentifier, elementIdsList, xiList):
    """
    Updates elementIdsList and xiList with the element ID and xi for each node in nIds
    :param nIds: List of node identifiers
    :param elementIdentifier: Element identifier for nodes in nIds
    :param elementIdsList: elementIdsList[node identifier] stores element identifier where node belongs to.
    :param xiList: xiList[node identifier] stores xi value of node along the element.
    return: updated elementIdsList and xiList
    """

    for n in range(len(nIds)):
        elementIdsList[nIds[n]] = elementIdentifier
        xiList[nIds[n]] = 1.0 if n else 0.0

    return elementIdsList, xiList